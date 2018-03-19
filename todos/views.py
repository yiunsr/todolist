import logging
from django.template.context_processors import csrf
import traceback
from django.template.response import TemplateResponse
from todolist.ErrClass import ErrClass
from rest_framework import viewsets
from todos.models import Todo
from todos.serializers import TodoSerializer, AdminTodoSerializer
from rest_framework.permissions import IsAuthenticated, IsAdminUser
import django_filters
from django.contrib.auth.decorators import login_required, user_passes_test
from rest_framework.decorators import list_route
import json
from django.http.response import HttpResponse
import csv
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
import mimetypes
import io
from openpyxl.reader.excel import load_workbook
import datetime
from util.utils import JSONResponse


logger = logging.getLogger('django_log')


class TodoViewSet(viewsets.ModelViewSet):
    """
    Account Model 을 Restful API 로 관리 하는 ViewSet
    """
    
    queryset = Todo.objects.all()  #pylint: disable=no-member
    serializer_class = TodoSerializer
    
    permission_classes = (IsAuthenticated  ,)
    
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend,)
    filter_fields  = ( 'id', 'priority')
    
    def get_queryset(self):
        user = self.request.user
        return Todo.objects.filter(user=user)  #pylint: disable=no-member
    
    @list_route(methods=['POST'])
    def save(self, request):
        """
        HTTP POST Method 를 사용하고  save 라는 path 를 사용할 때
        데이터를 저장하는 함수
        저장하려는 데이터와 실제 DB에 저장된 데이터가 있을 때, 
        그 차이에 대해서만 저장하고 삭제한다. 
        """
        try:
            jsonDataString = request.POST.get("jsonData")
            jsonDataList = json.loads(jsonDataString)
            josnDataDict = {}  ## id 를 key  로,  나머지 데이터를 value 로하는 dict
            index = 2**32
            for item in jsonDataList:
                _id = item["id"]
                if _id == "": ## 새로 추가 되는 아이템인 경우 id 가 0일 수 있는데, 이 경우 임의의 id를 부여
                    item["id"] = _id = index
                    index +=1
                josnDataDict[_id] = item
            jsonDataSet = set(josnDataDict.keys())
            
            objects = self.get_queryset().filter(user=request.user)
            objectsDict = {} ## id 를 key로, 나머지 데이터를 value 로하는 dict
            for item in objects:
                _id = item.id
                objectsDict[_id] = item
            objectsSet = set(objectsDict.keys())
            
            ## jsonData 와 objects(DB) 를 비교해서
            ## jsonData 에만 존재, 데이터 추가
            itemToAdd = jsonDataSet - objectsSet
            for _id in itemToAdd:
                item = josnDataDict[_id]
                create_time = datetime.datetime.fromtimestamp(item["create_time"]/1000)
                todo = Todo(user = request.user, priority = item["priority"], text = item["text"],  done = item["done"], create_time = create_time)
                todo.save()
            
            ## objects에만 존재 데이터 삭제
            itemToDelete = objectsSet - jsonDataSet
            for _id in itemToDelete:
                item = objectsDict[_id]
                self.get_queryset().filter( id = _id ).delete()
                
            ## 양쪽에 다 존재하는 데이터의 done과 priority 상태 확인
            itemToUpdateCheck = objectsSet.intersection(jsonDataSet)
            for _id in itemToUpdateCheck:
                itemWeb = josnDataDict[_id]
                itemDB = objectsDict[_id]
                if itemWeb["done"] != itemDB.done or itemWeb["priority"] != itemDB.priority:
                    self.get_queryset().filter( id = _id ).update( done = itemWeb["done"], priority = itemWeb["priority"]  )
            
            result = ErrClass('NOERROR').toDict()
            return HttpResponse(json.dumps(result), content_type="application/json")
        except Exception as e:
            logger.error(traceback.format_exc() )
            return ErrClass('UNKNWON_ERROR').response()
   
class AdminTodoViewSet(viewsets.ModelViewSet):
    """
    Account Model 을 Restful API 로 관리 하는 ViewSet
    """
    
    queryset = Todo.objects.all()  #pylint: disable=no-member
    serializer_class = AdminTodoSerializer
    
    ## admin 만 이 기능 사용가능
    permission_classes = (IsAdminUser  ,)
    
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend,)
    filter_fields  = ( 'id', 'priority', 'user' , "done")
    
    @list_route()
    def tablelist(self, request):
        """
        datatable.js 와 통신하는 함수
        """
        #TODO : 인젝션 방지를 위해 가능한 parameter 만 검사하는 코드 필요
        params = request.GET
         
        col_len = int(params.get('col_len', 0)) #전체 column 개수
        ord_col_len = int(params.get('ord_col_len', 0)) #정렬이 필요한 column 개수(현재는 1개만 가능) 
        start = int(params.get('start', 0)) # DB의 limit 에 해당하는 offset
        num = int(params.get('length', 25))  # 한 번에 가져와야 하는 리스트 개수
        week_day = params.get('extra[week_day]', "")
        
        column_list = []
        search_dict = {  }
        order_list = []
        
        # loop를 돌면서 column_list 정보와 필터링해야 하는 search_dict 정보를 가져온다. 
        for index in range(col_len):
            param_key = "columns[" +  str(index) + "][data]"
            col_name = params.get(param_key, None) 
            column_list.append( col_name )
            param_key = "columns[" +  str(index) + "][search][value]"
            search_value =  params.get(param_key, "")
            if  search_value:
                search_dict[col_name] = search_value
                
        
    
        # loop를 돌면서 정렬이 필요한 column 정보를 얻는다. 
        for index in range(ord_col_len):
            param_key = "order[" +  str(index) + "][column]"
            col_name = column_list[int(params.get(param_key, None))]
            order_dir_key = "order[" +  str(index) + "][dir]"
            order_dir = params.get(order_dir_key, None)
            if order_dir == "asc":
                order_list.append( col_name )
            else:
                order_list.append( "-" + col_name )
            
        
        obj_list = self.queryset
        if len(search_dict): #filtering 해야 하는 정보가 있으면 filter
            obj_list = obj_list.filter( **search_dict  )
        
        # extra 필터 정보 검색 
        if week_day:
            obj_list = obj_list.filter(create_time__week_day = int(week_day)  )
        
        obj_list = obj_list.order_by(*order_list) #데이터 정렬
        response_list = obj_list[start:(start+num)] # list 에 대한 limit start, start+num
        
        serializer = self.get_serializer(response_list, many=True)
  
        d = {
            "draw":params.get('draw',1),            
            "recordsTotal": obj_list.count(),
            "recordsFiltered": obj_list.count(), 
            "data" : serializer.data
        }
         
        return JSONResponse(self.request, d, status=200)
    

@login_required
def commonlist(request):
    try: 
        logger.info("/todos/commonlist")
        c = {}
        c.update(csrf(request))
        response = TemplateResponse(request, 'todos/list.html', c )
        response.render()
        return response
    
    except Exception as e:
        logger.error(traceback.format_exc() )
        return ErrClass('UNKNWON_ERROR').response()

@login_required
def dayofweeklist(request):
    try: 
        logger.info("/todos/dayofweeklist")
        c = {}
        c.update(csrf(request))
        if request.user.is_superuser:
            response = TemplateResponse(request, 'todos/admin_dayofweeklist.html', c )
        else:
            response = TemplateResponse(request, 'todos/dayofweeklist.html', c )
        response.render()
        return response
    
    except Exception as e:
        logger.error(traceback.format_exc() )
        return ErrClass('UNKNWON_ERROR').response()

@user_passes_test(lambda u: u.is_superuser)
def adminlist(request):
    try: 
        logger.info("/todos/adminlist")
        c = {}
        c.update(csrf(request))
        response = TemplateResponse(request, 'todos/adminlist.html', c )
        response.render()
        return response
    
    except Exception as e:
        logger.error(traceback.format_exc() )
        return ErrClass('UNKNWON_ERROR').response()  
    
@login_required
def exportData(request):
    try:
        filetype = request.GET.get("filetype")
        jsonDataString = request.GET.get("jsonData")
        jsonDataList = json.loads(jsonDataString)
        if filetype == "csv":
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="todolist.csv"'

            # id, priority, text, create_time, done
            writer = csv.writer(response)
            for item in jsonDataList:
                writer.writerow([item["id"], item["priority"], item["text"], item["create_time"], item["done"]] )
            return response
        elif filetype =="excel":
            wb = openpyxl.Workbook()
            worksheet = wb.active
            for item in jsonDataList:
                worksheet.append([item["id"], item["priority"], item["text"], item["create_time"], item["done"]])
            response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="todolist.xls"'
            return response
        elif filetype =="json":
            response = HttpResponse( content = jsonDataString, content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="todolist.json"'
            return response
        return ErrClass('UNKNWON_ERROR').response()
    
    except Exception as e:
        logger.error(traceback.format_exc() )
        return ErrClass('UNKNWON_ERROR').response()
    
@login_required
def importData(request):
    try:
        importfile = request.FILES["importfile"]
        file_mime = mimetypes.guess_type(importfile.name)
        jsonDataList = []
        importfile.seek(0)
        if file_mime[0] == "text/csv":
            reader = csv.reader(io.StringIO(importfile.read().decode('utf-8')))
            for row in reader:
                # id, priority, text, create_time, done
                done = row[4] == "True"
                item = {"id":row[0], "priority":row[1], "text":row[2], "create_time":int(row[3]), "done" : done}
                jsonDataList.append(item)
        elif file_mime[0] =="application/vnd.ms-excel":
            wb = load_workbook(filename=io.BytesIO(importfile.read()))
            worksheet = wb.active
            for row in worksheet.rows:
                # id, priority, text, create_time, done
                done = row[4].value
                item = {"id":row[0].value or "" , "priority": str(row[1].value), "text":row[2].value, "create_time":int(row[3].value), "done" : done}
                jsonDataList.append(item)
        elif file_mime[0] =="application/json":
            jsonDataByte = importfile.read()
            jsonDataStr = jsonDataByte.decode("utf-8") 
            jsonDataList =  json.loads(jsonDataStr)
        else:
            return ErrClass('UNKNWON_ERROR').response()
    
        result = ErrClass('NOERROR').toDict()
        result["jsonData"] = jsonDataList
        return HttpResponse(json.dumps(result), content_type="application/json")
    
    except Exception as e:
        logger.error(traceback.format_exc() )
        return ErrClass('UNKNWON_ERROR').response()
    
    